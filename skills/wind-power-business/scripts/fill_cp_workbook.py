#!/usr/bin/env python3
"""Fill the Cp calculation workbook template from a pasted power curve."""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency guidance path
    raise SystemExit(
        "openpyxl is required for xlsx output. Install with: "
        "pip install -r skills/wind-power-business/resources/requirements.txt"
    ) from exc

from compute_cp import calculate, parse_curve, read_input, validate_input_parameters, validate_physical_result

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_TEMPLATE = SKILL_DIR / "assets" / "cp-calculation-template.xlsx"
START_ROW = 9
END_ROW = 108


def set_recalculation(workbook) -> None:
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass


def fill_workbook(
    input_text: str,
    output_path: Path,
    template_path: Path,
    rho: float,
    diameter: float,
    allow_unphysical: bool = False,
    allow_suspicious_parameters: bool = False,
) -> None:
    validate_input_parameters(
        input_text,
        rho=rho,
        diameter_m=diameter,
        allow_suspicious_parameters=allow_suspicious_parameters,
    )
    rows = parse_curve(input_text)
    _, summary = calculate(rows, rho=rho, diameter_m=diameter)
    validate_physical_result(summary, input_text, allow_unphysical=allow_unphysical)

    capacity = END_ROW - START_ROW + 1
    if len(rows) > capacity:
        raise ValueError(f"Template supports {capacity} rows, got {len(rows)} rows.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Cp计算"] if "Cp计算" in workbook.sheetnames else workbook.active

    sheet["B4"] = rho
    sheet["B5"] = diameter
    sheet["E5"] = f"=MAX(E{START_ROW}:E{END_ROW})"

    for offset in range(capacity):
        row_index = START_ROW + offset
        if offset < len(rows):
            wind_speed, power, ct = rows[offset]
            sheet.cell(row_index, 1).value = wind_speed
            sheet.cell(row_index, 2).value = power
            sheet.cell(row_index, 3).value = ct
        else:
            sheet.cell(row_index, 1).value = None
            sheet.cell(row_index, 2).value = None
            sheet.cell(row_index, 3).value = None

        sheet.cell(row_index, 4).value = (
            f'=IF($A{row_index}="","",0.5*$B$4*$A{row_index}^3*PI()*$B$5^2/4/1000)'
        )
        sheet.cell(row_index, 5).value = (
            f'=IF(OR($A{row_index}="",$B{row_index}=""),"",$B{row_index}/$D{row_index})'
        )

    set_recalculation(workbook)
    workbook.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fill the Cp calculation workbook template from wind-speed, power, and optional Ct columns.",
    )
    parser.add_argument("input", nargs="?", default="-", help="Input text/CSV file. Use '-' or omit for stdin.")
    parser.add_argument("--rho", type=float, required=True, help="Required air density in kg/m^3.")
    parser.add_argument("--diameter", type=float, required=True, help="Required rotor diameter in meters.")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Cp workbook template path.")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    parser.add_argument(
        "--allow-unphysical",
        action="store_true",
        help="Allow xlsx output even when Cp exceeds the Betz limit. Use only for diagnostics.",
    )
    parser.add_argument(
        "--allow-suspicious-parameters",
        action="store_true",
        help="Allow xlsx output after the user confirms parameters flagged by broad sanity checks.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        fill_workbook(
            input_text=read_input(args.input),
            output_path=Path(args.output),
            template_path=Path(args.template),
            rho=args.rho,
            diameter=args.diameter,
            allow_unphysical=args.allow_unphysical,
            allow_suspicious_parameters=args.allow_suspicious_parameters,
        )
    except (OSError, ValueError) as exc:
        parser.error(str(exc))
        return 2

    return 0


if __name__ == "__main__":
    sys.exit(main())
