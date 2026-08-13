#!/usr/bin/env python3
"""Convert an XLSX/CSV/TSV coordinate table to a validated Shapefile bundle."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import os
import re
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable


def _load_dependencies() -> tuple[Any, Any, Any]:
    try:
        import shapefile
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency 'pyshp'. Install resources/requirements.txt first."
        ) from exc
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install resources/requirements.txt first."
        ) from exc
    try:
        import pyproj
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency 'pyproj'. Install resources/requirements.txt first."
        ) from exc
    return shapefile, openpyxl, pyproj


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def _read_table(path: Path, sheet_name: str | None, openpyxl: Any) -> tuple[list[str], list[list[Any]], str]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet {sheet_name!r} not found; available: {workbook.sheetnames}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_header = list(next(iterator))
        except StopIteration as exc:
            raise ValueError("The worksheet is empty") from exc
        rows = [list(row) for row in iterator if any(value not in (None, "") for value in row)]
        source_label = worksheet.title
        workbook.close()
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream, delimiter=delimiter)
            try:
                raw_header = list(next(reader))
            except StopIteration as exc:
                raise ValueError("The table is empty") from exc
            rows = [list(row) for row in reader if any(value != "" for value in row)]
        source_label = path.name
    else:
        raise ValueError("Supported input formats: .xlsx, .csv, .tsv")

    headers = [_text(value).strip() or f"FIELD_{index + 1:03d}" for index, value in enumerate(raw_header)]
    width = len(headers)
    normalized_rows = [(row + [None] * width)[:width] for row in rows]
    return headers, normalized_rows, source_label


def _resolve_column(spec: str | None, headers: list[str], candidates: set[str], label: str) -> int | None:
    if spec is None:
        lowered = [header.strip().lower() for header in headers]
        for index, header in enumerate(lowered):
            if header in candidates:
                return index
        return None
    if spec.isdigit():
        index = int(spec)
        if not 0 <= index < len(headers):
            raise ValueError(f"{label} column index {index} is out of range")
        return index
    lowered = [header.strip().lower() for header in headers]
    target = spec.strip().lower()
    if target not in lowered:
        raise ValueError(f"{label} column {spec!r} not found; headers: {headers}")
    return lowered.index(target)


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _safe_dbf_names(headers: list[str]) -> list[str]:
    names: list[str] = []
    used: set[str] = set()
    for index, header in enumerate(headers, start=1):
        ascii_name = re.sub(r"[^A-Za-z0-9_]", "_", header.encode("ascii", "ignore").decode())
        ascii_name = ascii_name.strip("_") or f"FIELD_{index:03d}"
        if ascii_name[0].isdigit():
            ascii_name = f"F_{ascii_name}"
        base = ascii_name[:10]
        candidate = base
        suffix_index = 2
        while candidate.upper() in used:
            suffix = f"_{suffix_index}"
            candidate = f"{base[:10 - len(suffix)]}{suffix}"
            suffix_index += 1
        used.add(candidate.upper())
        names.append(candidate)
    return names


def _infer_field(values: Iterable[Any]) -> tuple[str, int, int]:
    present = [value for value in values if value not in (None, "")]
    if present and all(isinstance(value, bool) for value in present):
        return "L", 1, 0
    if present and all(isinstance(value, (dt.date, dt.datetime)) for value in present):
        return "D", 8, 0
    if present and all(isinstance(value, int) and not isinstance(value, bool) for value in present):
        return "N", 18, 0
    if present and all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
        return "N", 20, 6
    width = max((len(_text(value).encode("utf-8")) for value in present), default=1)
    return "C", max(1, min(width, 254)), 0


def _dbf_value(value: Any, field_type: str) -> Any:
    if value in (None, ""):
        return None if field_type != "C" else ""
    if field_type == "C":
        return _text(value)
    if field_type == "D":
        if isinstance(value, dt.datetime):
            value = value.date()
        if isinstance(value, dt.date):
            return value
        text = _text(value)[:10].replace("-", "")
        return text if len(text) == 8 and text.isdigit() else None
    if field_type == "L":
        return bool(value)
    return _number(value)


def _local_wkt(name: str) -> str:
    safe_name = name.replace('"', "'")
    return (
        f'LOCAL_CS["{safe_name}",LOCAL_DATUM["Local_Datum",0],'
        'UNIT["metre",1.0],AXIS["Easting",EAST],AXIS["Northing",NORTH]]'
    )


def _write_bundle(
    output: Path,
    input_path: Path,
    headers: list[str],
    rows: list[list[Any]],
    source_label: str,
    x_index: int,
    y_index: int,
    z_index: int | None,
    prj_file: Path | None,
    local_crs_name: str | None,
    epsg: int | None,
    skip_invalid: bool,
    overwrite: bool,
    shapefile: Any,
    pyproj: Any,
) -> dict[str, Any]:
    if output.exists() and not overwrite:
        raise FileExistsError(f"Output exists: {output}; pass --overwrite to replace it")
    output.parent.mkdir(parents=True, exist_ok=True)

    dbf_names = _safe_dbf_names(headers)
    field_specs = [
        _infer_field(row[column] for row in rows)
        for column in range(len(headers))
    ]

    valid_rows: list[tuple[int, list[Any], float, float, float | None]] = []
    invalid_rows: list[int] = []
    for source_row, row in enumerate(rows, start=2):
        x = _number(row[x_index])
        y = _number(row[y_index])
        z = _number(row[z_index]) if z_index is not None else None
        if x is None or y is None or (z_index is not None and z is None):
            invalid_rows.append(source_row)
            if skip_invalid:
                continue
            continue
        valid_rows.append((source_row, row, x, y, z))

    if invalid_rows and not skip_invalid:
        raise ValueError(
            f"Invalid coordinate values in {len(invalid_rows)} rows: {invalid_rows[:20]}; "
            "fix them or pass --skip-invalid"
        )
    if not valid_rows:
        raise ValueError("No valid coordinate rows found")

    with tempfile.TemporaryDirectory(
        prefix=".tabular-to-shp-",
        dir=output.parent,
    ) as temp_name:
        temp_dir = Path(temp_name)
        base_name = re.sub(r"[^A-Za-z0-9_-]+", "_", input_path.stem).strip("_") or "coordinates"
        base = temp_dir / base_name
        shape_type = shapefile.POINTZ if z_index is not None else shapefile.POINT
        writer = shapefile.Writer(str(base), shapeType=shape_type, encoding="utf-8")
        writer.autoBalance = 1

        for dbf_name, (field_type, size, decimals) in zip(dbf_names, field_specs, strict=True):
            writer.field(dbf_name, field_type, size=size, decimal=decimals)

        for _, row, x, y, z in valid_rows:
            if z_index is None:
                writer.point(x, y)
            else:
                writer.pointz(x, y, z)
            record = [
                _dbf_value(value, field_specs[index][0])
                for index, value in enumerate(row)
            ]
            writer.record(*record)
        writer.close()

        base.with_suffix(".cpg").write_text("UTF-8\n", encoding="ascii")
        crs_kind = "unknown"
        crs_name = None
        crs_epsg = None
        if prj_file:
            prj_text = prj_file.read_text(encoding="utf-8")
            try:
                crs = pyproj.CRS.from_wkt(prj_text)
            except pyproj.exceptions.CRSError as exc:
                raise ValueError(f"Invalid PRJ WKT: {prj_file}") from exc
            base.with_suffix(".prj").write_text(prj_text, encoding="utf-8")
            crs_kind = "prj-file"
            crs_name = crs.name
            crs_epsg = crs.to_epsg()
        elif local_crs_name:
            base.with_suffix(".prj").write_text(_local_wkt(local_crs_name), encoding="utf-8")
            crs_kind = "local"
            crs_name = local_crs_name
        elif epsg is not None:
            try:
                crs = pyproj.CRS.from_epsg(epsg)
            except pyproj.exceptions.CRSError as exc:
                raise ValueError(f"Invalid or unsupported EPSG code: {epsg}") from exc
            base.with_suffix(".prj").write_text(
                crs.to_wkt(version="WKT1_ESRI"),
                encoding="utf-8",
            )
            crs_kind = "epsg"
            crs_name = crs.name
            crs_epsg = epsg

        mapping_path = temp_dir / f"{base_name}_field_mapping.csv"
        with mapping_path.open("w", encoding="utf-8-sig", newline="") as stream:
            mapping_writer = csv.writer(stream)
            mapping_writer.writerow(["source_index", "source_field", "dbf_field", "type", "width", "decimals"])
            for index, (header, dbf_name, spec) in enumerate(
                zip(headers, dbf_names, field_specs, strict=True)
            ):
                mapping_writer.writerow([index, header, dbf_name, *spec])

        reader = shapefile.Reader(str(base), encoding="utf-8")
        summary = {
            "source": str(input_path.resolve()),
            "sourceSheet": source_label,
            "records": len(reader),
            "skippedRows": invalid_rows if skip_invalid else [],
            "shapeType": reader.shapeTypeName,
            "bbox": [float(value) for value in reader.bbox],
            "zRange": (
                [float(value) for value in reader.zbox]
                if z_index is not None
                else None
            ),
            "fieldCount": len(reader.fields) - 1,
            "xColumn": headers[x_index],
            "yColumn": headers[y_index],
            "zColumn": headers[z_index] if z_index is not None else None,
            "crsKind": crs_kind,
            "crsName": crs_name,
            "crsEpsg": crs_epsg,
        }
        reader.close()
        report_path = temp_dir / f"{base_name}_conversion_report.json"
        report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        staged_output = temp_dir / output.name
        with zipfile.ZipFile(staged_output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for suffix in (".shp", ".shx", ".dbf", ".cpg", ".prj"):
                component = base.with_suffix(suffix)
                if component.exists():
                    archive.write(component, arcname=component.name)
            archive.write(mapping_path, arcname=mapping_path.name)
            archive.write(report_path, arcname=report_path.name)
        with zipfile.ZipFile(staged_output) as archive:
            bad_member = archive.testzip()
            if bad_member:
                raise RuntimeError(f"ZIP verification failed for member: {bad_member}")
        os.replace(staged_output, output)
    return summary


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert an XLSX/CSV/TSV coordinate table to a Point/PointZ Shapefile ZIP."
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--sheet", help="XLSX worksheet name; defaults to the active worksheet")
    parser.add_argument("--x-col", help="X/easting column name or zero-based index")
    parser.add_argument("--y-col", help="Y/northing column name or zero-based index")
    parser.add_argument("--z-col", help="Optional Z/elevation column name or zero-based index")
    crs_group = parser.add_mutually_exclusive_group(required=True)
    crs_group.add_argument("--prj-file", type=Path, help="Authoritative WKT .prj file to copy")
    crs_group.add_argument("--local-crs-name", help="Confirmed local coordinate-system name")
    crs_group.add_argument("--epsg", type=int, help="Authoritative EPSG code for the input coordinates")
    parser.add_argument("--skip-invalid", action="store_true", help="Skip rows with invalid X/Y/Z")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args(argv)

    try:
        shapefile, openpyxl, pyproj = _load_dependencies()
        input_path = args.input.resolve()
        if not input_path.is_file():
            raise ValueError(f"Input file not found: {input_path}")
        if args.prj_file and not args.prj_file.is_file():
            raise ValueError(f"PRJ file not found: {args.prj_file}")
        headers, rows, source_label = _read_table(input_path, args.sheet, openpyxl)
        x_index = _resolve_column(args.x_col, headers, {"x", "easting", "东坐标", "横坐标"}, "X")
        y_index = _resolve_column(args.y_col, headers, {"y", "northing", "北坐标", "纵坐标"}, "Y")
        z_index = _resolve_column(args.z_col, headers, {"z", "high", "height", "elevation", "高程"}, "Z")
        if x_index is None or y_index is None:
            raise ValueError("Unable to detect X/Y columns; pass --x-col and --y-col")
        output = args.output or input_path.with_name(f"{input_path.stem}.shapefile.zip")
        if output.suffix.lower() != ".zip":
            raise ValueError("Output must be a .zip file")
        summary = _write_bundle(
            output.resolve(),
            input_path,
            headers,
            rows,
            source_label,
            x_index,
            y_index,
            z_index,
            args.prj_file.resolve() if args.prj_file else None,
            args.local_crs_name,
            args.epsg,
            args.skip_invalid,
            args.overwrite,
            shapefile,
            pyproj,
        )
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"OUTPUT: {output.resolve()}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
